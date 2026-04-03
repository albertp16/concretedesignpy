# MIT License
# Copyright (c) Albert Pamonag Engineering Consultancy

"""Beam calculation API routes."""

import io
import math

from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from concretedesignpy.calculators.beam_moment import calculate_beam_moment
from concretedesignpy.provision import generate_phi_plot_data
from concretedesignpy.calculators.beam_shear import (
    compute_concrete_shear_strength,
    compute_steel_shear_strength,
    compute_shear_spacing,
    shear_torsion_design,
    shear_design,
)
from concretedesignpy.calculators.beam_torsion import torsion_design
from concretedesignpy.calculators.beam_deflection import deflection_computation
from concretedesignpy.calculators.diagrams import (
    svg_beam_cross_section,
    svg_shear_diagram,
    svg_torsion_section,
)

beam_bp = Blueprint("beam", __name__)


@beam_bp.route("/moment", methods=["POST"])
def beam_moment():
    """Calculate beam moment capacity."""
    data = request.get_json()
    try:
        result = calculate_beam_moment(
            rebar_list=data["rebar_list"],
            fc=float(data["fc"]),
            fy=float(data["fy"]),
            b=float(data["b"]),
            h=float(data["h"]),
            es=float(data.get("es", 200000)),
        )
        # Generate cross-section SVG
        result["svg"] = svg_beam_cross_section(
            b=float(data["b"]),
            h=float(data["h"]),
            rebar_forces=result["rebar_forces"],
            c=result["neutral_axis"],
            a=result["a"],
        )
        # Round rebar forces for display
        for rf in result["rebar_forces"]:
            rf["strain"] = round(rf["strain"], 6)
            rf["stress"] = round(rf["stress"], 2)
            rf["force"] = round(rf["force"] / 1000, 2)  # Convert to kN
            rf["area"] = round(rf["area"], 2)
        # Generate phi plot data
        result["phi_plot"] = generate_phi_plot_data(
            fy=float(data["fy"]),
            es=float(data.get("es", 200000)),
            actual_epsilon=result["epsilon_t"],
        )
        return jsonify({"status": "success", "result": result})
    except (KeyError, ValueError, TypeError) as e:
        return jsonify({"status": "error", "message": str(e)}), 400


@beam_bp.route("/shear", methods=["POST"])
def beam_shear():
    """Calculate beam shear capacity or required spacing."""
    data = request.get_json()
    try:
        mode = data.get("mode", "capacity")
        phi = float(data.get("phi", 0.75))

        if mode == "capacity":
            vc_result = compute_concrete_shear_strength(
                fc=float(data["fc"]),
                b=float(data["b"]),
                d=float(data["d"]),
                lamda=float(data.get("lamda", 1.0)),
                vc_type=data.get("vc_type", "simple"),
            )
            vs_result = compute_steel_shear_strength(
                av=float(data["av"]),
                fyt=float(data["fyt"]),
                d=float(data["d"]),
                s=float(data["s"]),
            )
            vn = vc_result["vc"] + vs_result["vs"]
            vu = phi * vn
            result = {
                **vc_result,
                **vs_result,
                "vn": round(vn, 2),
                "vn_kn": round(vn / 1000, 2),
                "vu": round(vu, 2),
                "vu_kn": round(vu / 1000, 2),
            }
            result["svg"] = svg_shear_diagram(
                vc_result["vc_kn"], vs_result["vs_kn"],
                round(vu / 1000, 2), phi,
            )
        else:
            result = compute_shear_spacing(
                fc=float(data["fc"]),
                b=float(data["b"]),
                d=float(data["d"]),
                fyt=float(data["fyt"]),
                vu_required=float(data["vu_required"]),
                phi=float(data.get("phi", 0.75)),
                av=float(data["av"]),
                lamda=float(data.get("lamda", 1.0)),
                vc_type=data.get("vc_type", "simple"),
            )

        return jsonify({"status": "success", "result": result})
    except (KeyError, ValueError, TypeError) as e:
        return jsonify({"status": "error", "message": str(e)}), 400


@beam_bp.route("/shear-torsion", methods=["POST"])
def beam_shear_torsion():
    """Combined shear and torsion design per ACI 318M-14."""
    data = request.get_json()
    try:
        result = shear_torsion_design(
            fc=float(data["fc"]),
            fyv=float(data["fyv"]),
            fy=float(data["fy"]),
            phi=float(data.get("phi", 0.75)),
            bw=float(data["bw"]),
            h=float(data["h"]),
            cc=float(data["cc"]),
            c=float(data["c"]),
            d=float(data["d"]),
            vu=float(data["vu"]),
            tu=float(data.get("tu", 0)),
            nu=float(data.get("nu", 0)),
            s_chosen=float(data.get("s_chosen", 150)),
            n_legs=int(data.get("n_legs", 4)),
            db_stirrup=float(data.get("db_stirrup", 10)),
            db_long=float(data.get("db_long", 12)),
        )
        return jsonify({"status": "success", "result": result})
    except (KeyError, ValueError, TypeError) as e:
        return jsonify({"status": "error", "message": str(e)}), 400


@beam_bp.route("/shear-design", methods=["POST"])
def beam_shear_design():
    """Shear-only design per ACI 318M-14."""
    data = request.get_json()
    try:
        result = shear_design(
            fc=float(data["fc"]),
            fyv=float(data["fyv"]),
            phi=float(data.get("phi", 0.75)),
            bw=float(data["bw"]),
            h=float(data["h"]),
            cc=float(data["cc"]),
            c=float(data["c"]),
            d=float(data["d"]),
            vu=float(data["vu"]),
            nu=float(data.get("nu", 0)),
            s_chosen=float(data.get("s_chosen", 150)),
            n_legs=int(data.get("n_legs", 4)),
            db_stirrup=float(data.get("db_stirrup", 10)),
        )
        return jsonify({"status": "success", "result": result})
    except (KeyError, ValueError, TypeError) as e:
        return jsonify({"status": "error", "message": str(e)}), 400


@beam_bp.route("/shear-design/export", methods=["POST"])
def beam_shear_export():
    """Export shear design report to Excel."""
    data = request.get_json()
    try:
        p = {k: float(data[k]) for k in ["fc", "fyv", "phi", "bw", "h", "cc", "c", "d", "vu"]}
        p["nu"] = float(data.get("nu", 0))
        p["s_chosen"] = float(data.get("s_chosen", 150))
        p["n_legs"] = int(data.get("n_legs", 4))
        p["db_stirrup"] = float(data.get("db_stirrup", 10))
        header = data.get("header", {})

        r = shear_design(**p)

        wb = Workbook()
        ws = wb.active
        ws.title = "BEAM SHEAR"
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # ── Styles ──
        hdr_font = Font(name="Arial", bold=True, size=11, color="1E3A8A")
        title_font = Font(name="Arial", bold=True, size=10, color="1E40AF")
        label_font = Font(name="Arial", size=9)
        value_font = Font(name="Arial", size=9, bold=True)
        unit_font = Font(name="Arial", size=9, color="6B7280")
        ref_font = Font(name="Arial", size=8, color="3B82F6", italic=True)
        safe_font = Font(name="Arial", size=9, bold=True, color="16A34A")
        unsafe_font = Font(name="Arial", size=9, bold=True, color="DC2626")
        hdr_fill = PatternFill("solid", fgColor="EFF6FF")
        section_fill = PatternFill("solid", fgColor="F0F9FF")
        result_fill = PatternFill("solid", fgColor="DCFCE7")
        thin_border = Border(
            bottom=Side(style="thin", color="E5E7EB"),
        )
        thick_border = Border(
            bottom=Side(style="medium", color="1E40AF"),
        )

        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 30

        row = 1

        # ── Header Block ──
        ws.merge_cells("B1:E1")
        ws["B1"] = header.get("company", "ENGINEERING CONSULTANT NAME")
        ws["B1"].font = Font(name="Arial", bold=True, size=14, color="1E3A8A")
        row = 2
        ws.merge_cells("B2:C2")
        ws["B2"] = header.get("address", "")
        ws["B2"].font = Font(name="Arial", size=8, color="6B7280")
        ws["D2"] = "Project:"
        ws["D2"].font = Font(name="Arial", size=8, bold=True)
        ws["E2"] = header.get("project", "")
        ws["E2"].font = Font(name="Arial", size=9, bold=True)
        row = 3
        ws["D3"] = "Job No:"
        ws["D3"].font = Font(name="Arial", size=8, bold=True)
        ws["E3"] = header.get("job_no", "")
        ws["E3"].font = Font(name="Arial", size=9)
        row = 4
        ws["B4"] = "Page Type:"
        ws["B4"].font = Font(name="Arial", size=8, bold=True)
        ws["C4"] = "CALCULATION SHEET"
        ws["C4"].font = Font(name="Arial", size=9, bold=True)
        ws["D4"] = "Prepared By:"
        ws["D4"].font = Font(name="Arial", size=8, bold=True)
        ws["E4"] = header.get("prepared_by", "")
        ws["E4"].font = Font(name="Arial", size=9)
        row = 5
        ws["B5"] = "Section:"
        ws["B5"].font = Font(name="Arial", size=8, bold=True)
        ws["C5"] = "STRENGTH DESIGN"
        ws["C5"].font = Font(name="Arial", size=9, bold=True)
        ws["D5"] = "Checked By:"
        ws["D5"].font = Font(name="Arial", size=8, bold=True)
        ws["E5"] = header.get("checked_by", "")
        ws["E5"].font = Font(name="Arial", size=9)
        for c_idx in range(2, 6):
            ws.cell(row=5, column=c_idx).border = thick_border

        row = 7
        ws.merge_cells("B7:E7")
        ws["B7"] = "DESIGN OF NONPRESTRESSED RECTANGULAR RC BEAMS FOR SHEAR - ACI 318M-14"
        ws["B7"].font = Font(name="Arial", bold=True, size=12, color="1E3A8A")
        ws["B7"].fill = hdr_fill

        def write_section(title, ref_text=""):
            nonlocal row
            row += 1
            ws.merge_cells(f"B{row}:D{row}")
            ws[f"B{row}"] = title
            ws[f"B{row}"].font = title_font
            ws[f"B{row}"].fill = section_fill
            if ref_text:
                ws[f"E{row}"] = ref_text
                ws[f"E{row}"].font = ref_font
            for c_idx in range(2, 6):
                ws.cell(row=row, column=c_idx).border = Border(
                    bottom=Side(style="thin", color="1E40AF"))
            return row

        def write_row(label, val, unit="", ref="", is_result=False, is_safe=None):
            nonlocal row
            row += 1
            ws[f"B{row}"] = label
            ws[f"B{row}"].font = label_font
            ws[f"C{row}"] = val
            ws[f"C{row}"].font = value_font
            ws[f"C{row}"].alignment = Alignment(horizontal="right")
            if is_result:
                ws[f"C{row}"].fill = result_fill
            if is_safe is True:
                ws[f"C{row}"].font = safe_font
            elif is_safe is False:
                ws[f"C{row}"].font = unsafe_font
            ws[f"D{row}"] = unit
            ws[f"D{row}"].font = unit_font
            if ref:
                ws[f"E{row}"] = ref
                ws[f"E{row}"].font = ref_font
            for c_idx in range(2, 6):
                ws.cell(row=row, column=c_idx).border = thin_border

        # ── Material Properties ──
        write_section("Material Properties", "ACI 318M-14 Cl. 20.2, 22.5.3")
        write_row("Cylinder Strength of Concrete, f'c =", p["fc"], "MPa",
                   "ACI 318M-14 Cl. 22.5.3.1")
        write_row("Yield Strength of Stirrups, fyv =", p["fyv"], "MPa",
                   "ACI 318M-14 Cl. 20.2.2.4")
        write_row("Strength Reduction Factor, \u03c6 =", p["phi"], "",
                   "ACI 318M-14 Cl. 21.2.1(b)")

        # ── Section Dimensions ──
        write_section("Section Dimensions & Concrete Covers")
        write_row("Width of Beam, bw =", p["bw"], "mm")
        write_row("Height of Beam, h =", p["h"], "mm")
        write_row("Clear Cover to Links, Cc =", p["cc"], "mm")
        write_row("Cover to Centroid of Reinforcement, c =", p["c"], "mm")
        write_row("Effective Depth of the Beam, d =", p["d"], "mm")

        # ── Straining Actions ──
        write_section("Straining Actions - Ultimate")
        write_row("Shear Force, Vu =", p["vu"], "kN")
        write_row("Axial Force, Nu =", p["nu"], "kN")

        # ── Concrete Shear Strength ──
        write_section("Computation of Concrete Shear Strength, Vc",
                       "ACI 318M-14 Cl. 22.5.5.1-7")
        write_row("Concrete Shear Strength, Vc =", r["vc"], "kN",
                   is_result=True)

        # ── Steel Shear Strength ──
        write_section("Computation of Transverse Reinforcement for Shear",
                       "ACI 318M-14 Cl. 22.5.10.5")
        write_row("Required Stirrup Shear Strength, Vs =", r["vs"], "kN")
        write_row("Max. Shear Strength by Stirrups, Vs,max =", r["vs_max"], "kN")
        is_safe = r["shear_status"] == "SAFE"
        write_row("Check Vu \u2264 \u03c6(Vc + Vs,max)", r["shear_status"], "",
                   "ACI 318M-14 Cl. 22.5.1.2", is_safe=is_safe)
        write_row("Required Shear Reinforcement, Av/s =", r["av_s_req"],
                   "mm\u00b2/mm", "ACI 318M-14 R22.5.10.5", is_result=True)

        # ── Minimum Reinforcement ──
        write_section("Minimum Shear Reinforcement",
                       "ACI 318M-14 Table 9.6.3.3")
        write_row("Minimum Stirrups Area, Av,min/s =", r["av_min"], "mm\u00b2/mm")
        write_row("Governing Av/s =", r["av_s_govern"], "mm\u00b2/mm",
                   is_result=True)

        # ── Chosen Reinforcement ──
        write_section("Chosen Reinforcement Check",
                       "ACI 318M-14 Cl. 9.7.6.2.2")
        write_row("Stirrup Spacing, s =", p["s_chosen"], "mm")
        write_row("Maximum Spacing, Smax =", r["smax"], "mm")
        sp_safe = r["spacing_ok"] == "OK"
        write_row("Spacing Check", r["spacing_ok"], "",
                   is_safe=sp_safe)
        write_row("No. of Stirrup Legs =", p["n_legs"], "")
        write_row("Stirrup Diameter =", p["db_stirrup"], "mm")
        write_row("Provided Av/s =", r["av_provided"], "mm\u00b2/mm")
        rft_safe = r["rft_ok"] == "OK"
        write_row("Reinforcement Check", r["rft_ok"], "",
                   is_safe=rft_safe, is_result=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Beam_Shear_Design_ACI318M14.xlsx",
        )
    except (KeyError, ValueError, TypeError) as e:
        return jsonify({"status": "error", "message": str(e)}), 400


@beam_bp.route("/torsion", methods=["POST"])
def beam_torsion():
    """Perform torsion design checks."""
    data = request.get_json()
    try:
        result = torsion_design(
            width=float(data["width"]),
            height=float(data["height"]),
            cover=float(data["cover"]),
            db=float(data["db"]),
            tf=float(data.get("tf", 0)),
            beff=float(data.get("beff", data["width"])),
            phi_torsion=float(data.get("phi_torsion", 0.75)),
            fc=float(data["fc"]),
            fy=float(data["fy"]),
            tu=float(data["tu"]),
            vc=float(data["vc"]),
            ds=float(data["ds"]),
            smax_shear=float(data["smax_shear"]),
            s_actual=float(data["s_actual"]),
            av=float(data["av"]),
            s=float(data["s"]),
        )
        g = result["geometry"]
        result["svg"] = svg_torsion_section(
            float(data["width"]), float(data["height"]),
            g["x1"], g["y1"], g["aoh"], g["ph"],
            float(data["cover"]), float(data["db"]),
        )
        return jsonify({"status": "success", "result": result})
    except (KeyError, ValueError, TypeError) as e:
        return jsonify({"status": "error", "message": str(e)}), 400


@beam_bp.route("/deflection", methods=["POST"])
def beam_deflection():
    """Calculate beam deflection."""
    data = request.get_json()
    try:
        result = deflection_computation(
            b=float(data["b"]),
            h=float(data["h"]),
            d=float(data["d"]),
            fc=float(data["fc"]),
            fy=float(data["fy"]),
            clearspan=float(data["clearspan"]),
            as_tension=float(data["as_tension"]),
            n_bars_comp=int(data.get("n_bars_comp", 0)),
            db_comp=float(data.get("db_comp", 0)),
            es=float(data.get("es", 200000)),
            ma=float(data["ma"]) if data.get("ma") else None,
            point_load=float(data.get("point_load", 0)),
            uniform_load=float(data.get("uniform_load", 0)),
            beam_type=data.get("beam_type", "simply_supported"),
            member_type=data.get("member_type", "floor"),
            sustained_duration=data.get("sustained_duration", "5_years_or_more"),
        )
        return jsonify({"status": "success", "result": result})
    except (KeyError, ValueError, TypeError) as e:
        return jsonify({"status": "error", "message": str(e)}), 400
